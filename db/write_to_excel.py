import os

from openpyxl import load_workbook

from functional.functions_tt import *

import pandas as pd
import pathlib


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            index_label='Индекс',
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, index_label='Индекс',
                **to_excel_kwargs)

    # save the workbook
    writer.save()


# TODO додеать чтоб ровно высчитывал длину значения в ячейке
def change_column_width(excel_file_name):
    writer = pd.ExcelWriter(excel_file_name, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(excel_file_name)

    for sheet in writer.book.worksheets:
        for col in sheet.columns:
            max_length = 0
            column_letter = col[0].column_letter  # Get the column name
            # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based)
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length * 1.1

            sheet.column_dimensions[column_letter].width = adjusted_width

    writer.save()


async def save_data_into_excel(excel_file_name):
    data = await get_data_all_tables()

    # удаляем файл с данными, если уже существует
    file = pathlib.Path(excel_file_name)
    if file.is_file():
        file.unlink()

    for table_info in data:
        # получаем значение первого ключа из словаря
        table_name = next(iter(table_info))
        table_data = table_info[table_name]

        excel_frame = pd.DataFrame(table_data[0], columns=table_data[1])
        append_df_to_excel(excel_file_name, excel_frame, sheet_name=table_name)

    change_column_width(excel_file_name)

# loop = asyncio.get_event_loop()
# loop.run_until_complete(save_data_into_excel())
